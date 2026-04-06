#!/usr/bin/env python3
"""
Excel (.xlsx) の各シートを UTF-8 BOM 付き CSV に書き出す。
  python3 xlsx_sheets_to_csv.py [入力.xlsx] [-o 出力ディレクトリ]

依存: pip install openpyxl
"""
from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

import openpyxl


def safe_sheet_filename(name: str) -> str:
    s = name.strip() or "sheet"
    s = re.sub(r'[\\/:*?"<>|]', "_", s)
    s = re.sub(r"\s+", "_", s)
    return s[:120] if len(s) > 120 else s


def cell_to_str(v):
    if v is None:
        return ""
    return v


def export_workbook(xlsx_path: Path, out_dir: Path, data_only: bool) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=data_only, read_only=True)
    out_dir.mkdir(parents=True, exist_ok=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out_path = out_dir / f"{safe_sheet_filename(sheet_name)}.csv"
        with out_path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, lineterminator="\n")
            for row in ws.iter_rows(values_only=True):
                w.writerow([cell_to_str(c) for c in row])
        print(out_path)

    wb.close()


def main() -> None:
    p = argparse.ArgumentParser(description="xlsx のシートごとに CSV を出力")
    p.add_argument(
        "xlsx",
        nargs="?",
        default=str(
            Path(__file__).resolve().parent.parent / "名称未設定フォルダ 2" / "1dunbine_data.xlsx"
        ),
        help="入力 .xlsx（省略時: dunbain/名称未設定フォルダ 2/1dunbine_data.xlsx）",
    )
    p.add_argument(
        "-o",
        "--out",
        default="",
        help="出力先ディレクトリ（省略時: 入力ファイルと同じ場所に <ファイル名>_csv）",
    )
    p.add_argument(
        "--formulas",
        action="store_true",
        help="計算結果ではなく式を読む（通常は使わない。data_only=False）",
    )
    args = p.parse_args()

    xlsx_path = Path(args.xlsx).resolve()
    if not xlsx_path.is_file():
        raise SystemExit(f"ファイルがありません: {xlsx_path}")

    if args.out:
        out_dir = Path(args.out).resolve()
    else:
        out_dir = xlsx_path.parent / f"{xlsx_path.stem}_csv"

    export_workbook(xlsx_path, out_dir, data_only=not args.formulas)


if __name__ == "__main__":
    main()
