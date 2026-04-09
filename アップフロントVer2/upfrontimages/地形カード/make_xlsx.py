"""
地形カード xlsx 作成スクリプト
各カードのダイス列(1,2b,3,4c,5o,6,7w,8,9,0r)の値と色(赤/黒)を記録。
値は "n" か "nR"(=赤)で指定。射程(レンジ)も "n" か "nR" で赤フラグを持つ。
実行するたびに /地形カード/terrain_cards.xlsx を上書き生成する。
途中保存用にも使う:読み終わったぶんだけ CARDS に追記していけばOK。
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- カードデータ ---
# (num, title, mod, range, subtype, [10 values], notes)
# range / value: "n" or "nR" (R = 赤字)
# subtype: FORD / FLANK / "" など
# notes: 特記(短文)
CARDS = [
    # 1-10
    (1,  "MOVEMENT", "+1", "6",  "",      ["1","1","1","1","1","1","1","1","1","1"], ""),
    (2,  "MOVEMENT", "+1", "5",  "",      ["2","2","2","2","2","2","2","2","2","2"], ""),
    (3,  "MOVEMENT", "+1", "5",  "FORD",  ["1R","1R","3","3","3","3","3","3","3","3"], ""),
    (4,  "WOODS",    "-2*","3",  "",      ["1R","2R","1R","4","4","4","4","4","4","4"], "NA to Mortar fire / INF from:1 shift left / AFV Entry/OVR BOG check / (RPC) Exit sideways / OVR OT AFV 6 IM / 5 EL Others 6 IM"),
    (5,  "MOVEMENT", "+1", "4",  "",      ["1R","1R","2R","1R","5","5","5","5","5","5"], ""),
    (6,  "MOVEMENT", "+1", "5",  "FLANK", ["1R","2R","3R","2R","1R","6","6","6","6","6"], ""),
    (7,  "MOVEMENT", "+1", "3",  "FORD",  ["1R","1","1","3","2R","1R","7","7","7","7"], ""),
    (8,  "MOVEMENT", "+1", "2",  "",      ["1R","2R","2R","4R","3R","2R","1R","8","8","8"], ""),
    (9,  "MOVEMENT", "+1", "2",  "",      ["1R","1R","3","1R","4R","3R","2R","1R","9","9"], ""),
    (10, "MOVEMENT", "+1", "1",  "",      ["1R","2R","1R","2R","5R","4R","3R","2R","1R","0"], ""),
    # 11-20
    (11, "MOVEMENT", "+1", "1",  "",      ["1R","2R","3","1R","5R","4R","3R","2R","1R","1"], ""),
    (12, "MOVEMENT", "+1", "0R", "",      ["1R","2R","3R","4R","2R","6","5R","4R","3R","2R"], ""),
    (13, "MOVEMENT", "+1", "0R", "FORD",  ["1R","1R","1R","3","1R","6R","5R","4R","3R","1R"], ""),
    (14, "MOVEMENT", "+1", "0R", "",      ["1R","2R","2R","2R","4R","2R","7R","6R","5R","4R"], ""),
    (15, "MOVEMENT", "+1", "0R", "FORD",  ["1R","1R","3R","3R","5R","3R","1R","7R","6R","5R"], ""),
    (16, "MOVEMENT", "+1", "1R", "",      ["1R","2R","1R","4R","1R","4R","2R","8R","7R","6R"], ""),
    (17, "MOVEMENT", "+1", "1R", "",      ["1R","1R","2R","1R","2R","5R","3R","1R","8R","7R"], ""),
    (18, "MOVEMENT", "+1", "4",  "FLANK", ["1R","2R","3R","2R","3R","6R","4R","2R","9R","8R"], ""),
    (19, "MOVEMENT", "+1", "2R", "",      ["1R","1R","1R","3R","4R","1R","5R","3R","1R","9R"], ""),
    (20, "MOVEMENT", "+1", "2R", "",      ["1R","2R","2R","4R","5R","2R","6R","4R","2R","0R"], ""),
    # 21-30
    (21, "MOVEMENT", "+1", "3",  "FORD",  ["1R","1R","3R","1R","3R","7","5R","3R","1R","0R"], ""),
    (22, "WOODS",    "-2*","1",  "",      ["1R","2R","1R","2R","2R","4R","1R","6R","4R","2R"], "NA to Mortar fire / INF from:1 shift left / AFV Entry/OVR BOG check / (RPC) Exit sideways"),
    (23, "MOVEMENT", "+1", "4R", "FLANK", ["1R","1R","2R","3R","3R","5R","2R","7R","5R","3R"], ""),
    (24, "MOVEMENT", "+1", "4R", "",      ["1R","2R","3R","4R","4R","6R","3R","8R","6R","4R"], ""),
    (25, "WOODS",    "-2*","0R", "",      ["1R","1R","1R","5R","1R","4R","1R","7R","5R","3R"], "NA to Mortar fire / INF from:1 shift left / AFV Entry/OVR BOG check / (RPC) Exit sideways"),
    (26, "MOVEMENT", "+1", "5R", "",      ["1R","2R","2R","2R","1R","2R","5R","2R","8R","6R"], ""),
    (27, "MOVEMENT", "+1", "5R", "FORD",  ["1R","1R","3R","3R","2R","3R","6R","3R","9R","7R"], ""),
    (28, "MOVEMENT", "+1", "6R", "",      ["1R","2R","1R","4R","3R","4R","7R","4R","1R","8R"], ""),
    (29, "MOVEMENT", "+1", "5R", "FLANK", ["1R","1R","2R","1R","4R","5R","1R","5R","2R","9R"], ""),
    (30, "HERO",     "",   "0R", "",      ["1R","2R","3R","2R","5R","6","2R","6R","3R","0R"], "Rally 1 Anytime; No Action / Doubles Firepower or +1 To Hit Frequency / Negates Wound Effects for one turn / Vs OVR Reduce RNC by one; Negate AFV Pin"),
    # 31-40
    (31, "RALLY 1",  "",   "6R", "BREEZE",["1","1","1","3R","1","1","3R","7R","4","1"], "Limit: one Rally card per Group per turn / BREEZE removes all Smoke after any RNC/RPC draw"),
    (32, "RALLY 1",  "",   "6",  "RADIO", ["1","2","2","4R","2","2","4R","2","2","4R"], "RADIO requires unpinned leader & RNC: No Rally"),
    (33, "BUILDINGS","-2", "1",  "",      ["1R","1","3","1","3","3","5","1R","6R","3"], "No outgoing mortar fire / OVR OT AFV 6 IM / 5 EL Others 6 IM"),
    (34, "RALLY 1",  "",   "5",  "",      ["1R","2R","1R","2R","4","4","6R","2R","7R","4"], "Partial credit allowed on stunned/pinned AFV"),
    (35, "RALLY 2",  "",   "1",  "BREEZE",["1R","1R","2R","3","5","5","7","3R","8R","5"], "BREEZE removes all Smoke after any RNC/RPC draw"),
    (36, "RALLY 5",  "",   "1R", "RADIO", ["1","2","3R","4R","1","6R","1","4R","9R","6"], "RADIO requires unpinned leader & RNC: No Rally"),
    (37, "RALLY 1",  "",   "4",  "",      ["1","1","1","2R","1","2R","1","2R","5R","1R"], "Partial credit allowed on stunned/pinned AFV"),
    (38, "RALLY 2",  "",   "2",  "BREEZE",["1R","2R","2R","2R","3R","2R","3R","6R","2R","8"], "BREEZE removes all Smoke after any RNC/RPC draw"),
    (39, "RALLY 6",  "",   "1",  "RADIO", ["1R","1R","3R","3R","4R","3R","4R","7","3R","9"], "RADIO requires unpinned leader & RNC: No Rally"),
    (40, "RALLY 1",  "",   "5R", "BREEZE",["1R","2R","1R","4R","5R","4R","5R","8","4R","0"], "BREEZE removes all Smoke after any RNC/RPC draw"),
    # 41-50
    (41, "RALLY 1",  "",   "4R", "",      ["1R","1R","2R","1R","1R","5R","6R","5R","1R","1R"], "Partial credit allowed on stunned/pinned AFV"),
    (42, "RALLY 2",  "",   "0R", "RADIO", ["1R","2R","3R","2R","2R","6R","7R","6R","2R","2R"], "RADIO requires unpinned leader & RNC: No Rally"),
    (43, "RALLY 3",  "",   "2",  "BREEZE",["1R","1R","1R","3R","3R","1R","1R","3R","7R","3"], "BREEZE removes all Smoke after any RNC/RPC draw"),
    (44, "RALLY 3",  "",   "2",  "BREEZE",["1R","2R","2R","4R","4R","2R","2R","4R","8","4"], "BREEZE removes all Smoke after any RNC/RPC draw"),
    (45, "BUILDINGS","-3", "1",  "",      ["1R","1R","3R","1R","5","3","3","5","9","5"], "No outgoing mortar fire / OVR OT AFV 6 IM / 5 EL Others 6 IM"),
    (46, "RALLY 3",  "",   "0R", "RADIO", ["1R","2R","1R","2R","1R","4","4","6R","1R","6"], "RADIO requires unpinned leader & RNC: No Rally"),
    (47, "RALLY 4",  "",   "3",  "",      ["1R","1R","2R","3R","2R","5","5","7R","2R","7"], "Partial credit allowed on stunned/pinned AFV"),
    (48, "RALLY 5",  "",   "1",  "BREEZE",["1R","2R","3R","4R","3R","6","6","8","3R","8"], "BREEZE removes all Smoke after any RNC/RPC draw"),
    (49, "RALLY 2",  "",   "2",  "BREEZE",["1R","1R","1R","1R","4R","1","7","1","4R","9"], "BREEZE removes all Smoke after any RNC/RPC draw"),
    (50, "RALLY 4",  "",   "3",  "BREEZE",["1R","2R","2R","2R","5R","2R","1R","2R","5R","0"], "BREEZE removes all Smoke after any RNC/RPC draw"),
    # 51-60
    (51, "RALLY 2",  "-1", "1",  "COWER/RADIO",["1R","1","3","3","1R","3","1R","3","6","1"], "COWER / RADIO: requires unpinned leader & RNC: No Rally"),
    (52, "RALLY 3",  "-2", "0R", "COWER/RADIO",["1R","2R","1R","4","2R","4","2R","4","3","4R"], "COWER / RADIO: requires unpinned leader & RNC: No Rally"),
    (53, "RALLY 4",  "-1", "4",  "COWER/RADIO",["1R","1R","2R","1R","3R","5","4","5","8","3"], "COWER / RADIO: requires unpinned leader & RNC: No Rally"),
    (54, "CONCEALED","-1", "3",  "",      ["1R","2R","3R","2R","4","6","5","6","9","4"], "NA AFV, Pillbox, Minefield, Sniper, fired/moved IG / CC: alters opponent's CCV / INF 1 shift left / Vs Ordnance: -1 To Hit Frequency; not Hit Effect"),
    (55, "CONCEALED","-1", "2",  "",      ["1R","1","1","3","5","1","6","7","1","5"], "NA AFV, Pillbox, Minefield, Sniper, fired/moved IG / Vs Ordnance: -1 To Hit Frequency"),
    (56, "CONCEALED","-1", "1",  "",      ["1R","2","2","4","1R","2","7","8","2","6"], "NA AFV, Pillbox, Minefield, Sniper, fired/moved IG"),
    (57, "CONCEALED","-1", "0R", "",      ["1R","1","3","1","2R","3","1","3","3","7"], "NA AFV, Pillbox, Minefield, fired/moved IG (heavy concealment) / INF 2 shifts left"),
    (58, "BUILDINGS","-2", "0R", "",      ["1R","2R","1R","2","3","4","2R","2","4","8"], "No outgoing mortar fire / OVR OT AFV 6 IM / 5 EL Others 6 IM"),
    (59, "CONCEALED","-1", "1R", "",      ["1R","1R","2R","3","4","5","3","3","5","9"], "NA AFV, Pillbox, Minefield, Sniper, fired/moved IG"),
    (60, "CONCEALED","-1", "2R", "",      ["1R","2R","3R","4","5","6","4","4","6","0"], "NA AFV, Pillbox, Minefield, fired/moved IG (heavy concealment) / INF 2 shifts left"),
    # 61-70
    (61, "CONCEALED","-1", "3",  "",      ["1","1","1","1","1","1","5","7","1","1"], "NA AFV, Pillbox, Minefield, Sniper, fired/moved IG"),
    (62, "CONCEALED","-1", "0R", "",      ["1","2","2","2","2","2","6","8","2","2"], "NA AFV, Pillbox, Minefield, Sniper, fired/moved IG"),
    (63, "CONCEALED","-2", "4",  "",      ["1R","1","3","3","3","3","7","9","3","3"], "NA AFV, Pillbox, Minefield, Sniper, fired/moved IG / INF 2 shift left / Vs Ord -2 To Hit Freq"),
    (64, "PILLBOX",  "-4", "0R", "Group B",["1R","2R","1R","4","4","4","1R","8","1R","4"], "3 men maximum; INF into 1 shift right / NA vs Flank & Elevated Fire, OVR, CC, Wire / NA from mortar, LATW, IG, AFV, SW, CC rear fire"),
    (65, "CONCEALED","-2", "2",  "",      ["1R","1R","2R","1R","5","5","2R","1R","2R","5"], "NA AFV, Pillbox, Minefield, fired/moved IG (heavy concealment) / INF 2 shifts left"),
    (66, "CONCEALED","-2", "2",  "",      ["1R","2R","3R","2R","1R","6","3","2R","3","6"], "NA AFV, Pillbox, Minefield, Sniper, fired/moved IG"),
    (67, "CONCEALED","-2", "4",  "",      ["1R","1R","1R","3R","2R","1R","4","3","4","7"], "NA AFV, Pillbox, Minefield, Sniper, fired/moved IG"),
    (68, "CONCEALED","-3", "1",  "",      ["1R","2R","2R","4R","3R","2R","5","4","5","8"], "NA AFV, Pillbox, Minefield, Sniper, fired/moved IG / INF 3 shift left / Vs Ord -3 To Hit Freq"),
    (69, "CONCEALED","-3", "1",  "",      ["1R","1R","3R","1R","4R","3R","6","5","6","9"], "NA AFV, Pillbox, Minefield, Sniper, fired/moved IG / INF 3 shift left"),
    (70, "FIRE 1",   "",   "6",  "FP①",   ["1R","2R","1R","2R","5R","4R","7","6","7","0"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    # 71-80
    (71, "BUILDINGS","-3", "0R", "",      ["1R","1R","2R","3","1R","5","1R","7","8","1R"], "No outgoing mortar fire / OVR OT AFV 6 IM / 5 EL Others 6 IM"),
    (72, "FIRE 2",   "",   "5",  "FP③",   ["1R","2R","3R","4","2R","6","2","8","9","2"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (73, "FIRE 3",   "",   "4",  "FP⑦",   ["1R","1R","1R","3","1R","3","1R","3","1R","2"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (74, "FIRE 1",   "",   "3",  "FP③",   ["1R","2R","2R","4","2R","4","2R","4","2R","4"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (75, "FIRE 2",   "",   "2",  "FP④",   ["1R","1R","3R","3","5","3","5","3","3","5"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (76, "FIRE 3",   "",   "4",  "FP⑥",   ["1R","2R","1R","4","1","4","6","4","6","7"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (77, "BUILDINGS","-2", "0R", "",      ["1R","1R","2R","1","2","5","7","5","5","5"], "No outgoing mortar fire / OVR OT AFV 6 IM / 5 EL Others 6 IM"),
    (78, "FIRE 1",   "",   "3",  "FP④",   ["1R","2R","3R","2","3","6","1","6","6","8"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (79, "FIRE 2",   "",   "1",  "FP⑥",   ["1R","1R","1R","3","4R","1","2","7","7","1"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (80, "FIRE 3",   "",   "1",  "FP⑧",   ["1R","2R","2R","4","5","2","3","8","8","0"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    # 81-90
    (81, "FIRE 1",   "",   "2",  "FP⑤",   ["1R","1","1","1","3","4","1","9","1","1"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (82, "FIRE 2",   "",   "1",  "FP⑦",   ["1R","2R","1","2","2","4","5","2","1R","2"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (83, "FIRE 3",   "",   "1R", "FP⑨",   ["1R","1R","2","3","3","5","6","3","2","3"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (84, "FIRE 1",   "",   "1R", "FP②",   ["1R","2R","3","4","4","6","7","4","3","4"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (85, "FIRE 2",   "",   "1R", "FP⑧",   ["1R","1R","1","1","5","1","5","1","4","5"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (86, "FIRE 3",   "",   "2R", "FP③",   ["1R","2R","2","2","1","2","2","6","5","6"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (87, "FIRE 1",   "",   "1R", "FP②",   ["1R","1R","3","3","2","3","3","7","6","7"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (88, "FIRE 2",   "",   "1R", "FP②",   ["1R","2R","1","4","3","4","4","8","7","8"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (89, "FIRE 3",   "",   "4",  "FP⑤",   ["1R","1R","2","1","4","5","5","1","8","9"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (90, "FIRE 4",   "",   "3",  "FP⑩",   ["1R","2R","3R","2","5R","6","6","2","9","0"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    # 91-100
    (91, "FIRE 4",   "",   "1",  "FP⑪",   ["1","1","1","3R","1","1","7","3","1","1"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (92, "FIRE 4",   "",   "0R", "FP⑧",   ["1","2","2","4R","2","2","1","4","2","2"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (93, "FIRE 4",   "",   "0R", "FP⑨",   ["1R","1","3","1","3","3","2","5","3","3"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (94, "FIRE 4",   "",   "1R", "FP⑫",   ["1R","2R","1","2","4","4","3","6","4","4"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (95, "BUILDINGS","-3", "0R", "",      ["1R","1R","2","3","5","5","4","7","5","5"], "No outgoing mortar fire / OVR OT AFV 6 IM / 5 EL Others 6 IM"),
    (96, "FIRE 4",   "",   "3",  "FP⑦",   ["1R","2R","3","4","1R","6","5","8","6","6"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (97, "FIRE 5",   "",   "2",  "FP⑬",   ["1R","1R","1","2","1R","6","1","7","7","7"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (98, "BUILDINGS","-2", "1",  "",      ["1R","2R","2","3","2R","7","2","8","8","8"], "No outgoing mortar fire / OVR OT AFV 6 IM / 5 EL Others 6 IM"),
    (99, "FIRE 5",   "",   "1",  "FP⑫",   ["1R","1R","3","4R","3R","1","3","9","9","9"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (100,"FIRE 5",   "",   "0R", "FP⑭",   ["1R","2R","1R","4R","5R","2","4","1","0","0"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    # 101-110
    (101,"FIRE 5",   "",   "0R", "FP⑪",   ["1R","2R","1R","1","1","5R","3","5","2","1"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (102,"FIRE 5",   "",   "1R", "FP⑮",   ["1R","2R","3R","2","2","6R","4","6","3","2"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (103,"FIRE 5",   "",   "2R", "FP⑩",   ["1R","1R","1R","3","3","1","5","3","2","3"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (104,"FIRE 6",   "",   "2R", "FP⑯",   ["1R","2R","2R","4","4","2","6","8","5","4"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (105,"FIRE 6",   "",   "1R", "FP⑰",   ["1R","1R","3R","1","5","3","7","1","6","5"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (106,"FIRE 6",   "",   "0R", "FP⑮",   ["1R","2R","1R","2","1","4","1","2","7","6"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (107,"FIRE 6",   "",   "0R", "FP⑬",   ["1R","1R","2R","3","2","5","2","3","8","7"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (108,"FIRE 6",   "",   "1R", "FP⑭",   ["1R","2R","3R","4","3","6","3","4","9","8"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (109,"FIRE 6",   "",   "2R", "FP⑱",   ["1R","1R","1R","1","4R","1","4","5","1","9"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (110,"FIRE 7",   "",   "3",  "FP⑯",   ["1R","2R","2R","2","5R","2","5","2","6","2"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    # 111-120
    (111,"FIRE 7",   "",   "0R", "FP⑮",   ["1R","1","3","3","1","3","6","3","7","3"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (112,"FIRE 7",   "",   "0R", "FP⑰",   ["1R","2","1","4","2","4","7","8","4","2"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (113,"FIRE 7",   "",   "3",  "FP⑱",   ["1R","1R","2","1","3","5","1","1","5","3"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (114,"FIRE 8",   "",   "2",  "FP⑰",   ["1R","2R","3","2","4","6","2","6","6","4"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (115,"FIRE 8",   "",   "2",  "FP⑱",   ["1R","1R","1","3","5","1","3","3","7","5"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (116,"FIRE 1",   "",   "3",  "FP③",   ["1R","2R","2","4R","1","2","4","4","8","6"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (117,"FIRE 1",   "",   "3",  "FP④",   ["1R","1R","3","1","2R","3","5","5","9","7"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (118,"GULLY",    "NA/-2*","3","",    ["1R","2R","1R","2","3","4","6","6","7","8"], "*Only from Hill, Mortar, or RR 5 / EXIT 2 Moves, or Move & Terrain / AFV Entry/OVR BOG check (RPC)"),
    (119,"WALL",     "-2*/-1","2","",    ["1R","1R","2","3","4R","5","7","7","2R","9"], "Vs opposite group only: No TEM if encircled / OVR OT AFV 6 IM / 5 EL Others 6 IM / AFV/IG Hull Down unless flanked or moving"),
    (120,"WIRE",     "⊙-1/+1","2","",   ["1R","2R","3R","4","5","6","1","8","3R","0"], "NA: Rejection, Marsh, Pillbox, unforded Stream / non-moving fully crewed weapon: PF +1 To Hit / NA INF, Flank Fire, Entrenching / AFV Entry/OVR BOG check (RPC) / Remove with Move if unpinned"),
    # 121-130
    (121,"GULLY",    "NA/-2*","3","",   ["1","1","1","1R","1","1","2","1","4","1"], "*Only from Hill, Mortar, or RR 5 / EXIT 2 Moves, or Move & Terrain / AFV Entry/OVR BOG check (RPC)"),
    (122,"WALL",     "-2*/-1","0R","",  ["1","2","2","2R","2","2","3","2","5","2"], "Vs opposite group only: No TEM if encircled / OVR OT AFV 6 IM / 5 EL Others 6 IM / AFV/IG Hull Down unless flanked or moving"),
    (123,"WIRE",     "⊙-1/+1","0R","",  ["1R","1","3","3","3","3","4","3","6","3"], "NA: Rejection, Marsh, Pillbox, unforded Stream / non-moving fully crewed weapon: PF +1 To Hit / NA INF, Flank Fire, Entrenching / Remove with Move if unpinned"),
    (124,"GULLY",    "NA/-2*","4","",   ["1R","2R","1R","4","4","4","5","4","7","4"], "*Only from Hill, Mortar, or RR 5 / EXIT 2 Moves, or Move & Terrain / AFV Entry/OVR BOG check (RPC)"),
    (125,"WALL",     "-2*/-1","2","",   ["1R","1R","2R","1","5","5","6","5","8","5"], "Vs opposite group only: No TEM if encircled / AFV/IG Hull Down unless flanked or moving"),
    (126,"WIRE",     "⊙-1/+1","2","",   ["1R","2R","3R","2","1R","6","7","6","9","6"], "NA: Rejection, Marsh, Pillbox, unforded Stream / Remove with Move if unpinned"),
    (127,"BRUSH",    "-1",   "1",  "",   ["1R","1R","1R","3","2R","1R","1","7","1","7"], "OVR OT AFV 6 IM / 5 EL Others 6 IM / INF from: 2 shifts left"),
    (128,"SMOKE",    "⊙-1/-1","3","",   ["1","2","2","4","3R","2","8","2","8","2"], "NA in Marsh: Unpinned SL,ASL,AFV in Group / INF from/into 2 shifts left per Smoke card / OVR OT AFV 6 IM / 5 EL Others 6 IM"),
    (129,"BRUSH",    "-1",   "0R", "",   ["1R","1","3","1","4R","3","1","3","1","9"], "OVR OT AFV 6 IM / 5 EL Others 6 IM / INF from: 2 shifts left"),
    (130,"SMOKE",    "⊙-1/-1","1","",   ["1R","2","1R","2","5","4","2","4","2","4"], "NA in Marsh: Unpinned SL,ASL,AFV in Group / INF from/into 2 shifts left per Smoke card"),
    # 131-140
    (131,"BRUSH",    "-1",   "0R", "",   ["1R","1R","2R","3","1R","5","5","3","3","1"], "OVR OT AFV 6 IM / 5 EL Others 6 IM / INF from: 2 shifts left"),
    (132,"SMOKE",    "⊙-1/-1","1","",   ["1R","2R","3R","4","2R","6","6","4","6","2"], "NA in Marsh: Unpinned SL,ASL,AFV in Group / INF from/into 2 shifts left per Smoke card"),
    (133,"BRUSH",    "-1",   "0R", "",   ["1R","1","1","1","3","1","7","5","7","3"], "OVR OT AFV 6 IM / 5 EL Others 6 IM / INF from: 2 shifts left"),
    (134,"SMOKE",    "⊙-1/-1","3","",   ["1R","2","2","2","4R","2","1","6","8","4"], "NA in Marsh: Unpinned SL,ASL,AFV in Group / INF from/into 2 shifts left per Smoke card"),
    (135,"BRUSH",    "-1",   "0R", "",   ["1R","1","3","3","5R","3","2","7","9","5"], "OVR OT AFV 6 IM / 5 EL Others 6 IM / INF from: 2 shifts left"),
    (136,"STREAM",   "⊙-1/0","0R", "",  ["1R","2R","1R","4","1","4","3","8","1","6"], "Rejection NA: AFV Entry/OVR BOG check (RPC) / No mortar, IG, MMG fire: PF To Hit -1 (all other non-moving fully crewed weapon)"),
    (137,"MARSH",    "⊙-1/-1","0R","",  ["1R","1R","2R","1","2","5","4","1","2","7"], "EXIT 1 sideways Move plus any other Move / No AFV/IG entry: No mortar, MMG may fire (all other non-moving fully crewed weapons)"),
    (138,"MINEFIELD","Atk:4 [1]","0R","",["1R","2R","3R","2","3","6","5","2","3","8"], "Rejection NA: attacks RPC with RNC (no color) / Exit 2 side Moves: RNC > 4 causes attack / Removal: RNC >= 4: RNC > 4 causes attack"),
    (139,"STREAM",   "⊙-1/0","0R", "",  ["1R","1R","1R","3","4","1","6","3","4","9"], "Rejection NA: AFV Entry/OVR BOG check (RPC) / No mortar, IG, MMG fire: PF To Hit -1"),
    (140,"MARSH",    "⊙-1/-1","0R","",  ["1R","2R","2R","4","5","2","7","4","5","0"], "EXIT 1 sideways Move plus any other Move / No AFV/IG entry: No mortar, MMG may fire"),
    # 141-150
    (141,"MINEFIELD","Atk:5 [1]","0R","",["1R","1","3","1","1","3R","1","5","6","1"], "Rejection NA: attacks RPC with RNC (no color) / Exit 2 side Moves: RNC > 3 causes attack / Removal: RNC >= 5: RNC > 4 causes attack"),
    (142,"HILL",     "+1*/-1*","2","",  ["1R","2R","1R","2","2","4","2","6","7","2"], "*Only if target/firing group is not also on Hill / INF from Hill: 1 shift right / AFV/IG Hull Down unless moving/flanked: No TEM"),
    (143,"BUILDINGS","-3", "1",  "",     ["1R","1R","2R","3","3","5","3","7","8","3"], "No outgoing mortar fire / OVR OT AFV 6 IM / 5 EL Others 6 IM"),
    (144,"SNIPER",   "KIA:5-6, PIN:3-4","0R","",["1R","2R","3R","4","4","6","4","8","9","4"], "Play as Discard only: ignore color of RNC / RPC determines target in multi card group / SNIPER CHECK: Action: RNC > Sniper RNC"),
    (145,"HILL",     "+1*/-1*","0R","", ["1R","1R","1R","1","5","1","5","1","1","5"], "*Only if target/firing group is not also on Hill / INF from Hill: 1 shift right / AFV/IG Hull Down unless moving/flanked: No TEM"),
    (146,"WOODS",    "-2*","1",  "",     ["1R","2R","2R","1","2","6","2","2","2","6"], "*NA to Mortar fire: INF from 1 shift left / AFV Entry/OVR BOG check (RPC) Exit sideways / OVR OT AFV 6 IM / 5 EL Others 6 IM"),
    (147,"WOODS",    "-2*","3",  "",     ["1R","1R","3R","2","3","7","3","3","3","7"], "*NA to Mortar fire: INF from 1 shift left / AFV Entry/OVR BOG check (RPC) Exit sideways"),
    (148,"SNIPER",   "KIA:6, PIN:4-5","0R","",["1R","2R","1R","4","3","4","1","4","4","8"], "Play as Discard only: ignore color of RNC / RPC determines target in multi card group / SNIPER CHECK: Action: RNC > Sniper RNC"),
    (149,"HILL",     "+1*/-1*","2","",  ["1R","1R","2R","1","4","5","2","5","5","9"], "*Only if target/firing group is not also on Hill / INF from Hill: 1 shift right / AFV/IG Hull Down unless moving/flanked: No TEM"),
    (150,"SNIPER",   "KIA:4-6, PIN:2-3","0R","",["1R","2R","3R","2","5","6","3","6","6","0"], "Play as Discard only: ignore color of RNC / RPC determines target in multi card group / SNIPER CHECK: Action: RNC > Sniper RNC"),
    # 151-162
    (151,"HERO",     "Rally 1 Anytime; No Action","0R","",["1R","1R","1R","3","1R","1R","4","7","1R","1R"], "Doubles Firepower or +1 To Hit Frequency / Negates Wound Effects for one turn / Vs OVR Reduce RNC by one; Negate AFV Pin"),
    (152,"FIRE 1",   "",   "6R", "FP①",   ["1R","2R","2R","4","2","2","5R","8","8","2"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (153,"FIRE 2",   "",   "2",  "FP⑤",   ["1R","1R","3","1","3","3","6","1","9","3"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (154,"FIRE 2",   "",   "5R", "FP①",   ["1R","2R","1","2","4","4","7","2","1R","4"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (155,"FIRE 3",   "",   "4R", "FP④",   ["1R","1R","2","3","5","5","1R","3","2R","5"], "Half Fire Strength: Moving Fire, Night Fire / Ordnance may use any Fire card for To Hit attempt"),
    (156,"SNIPER",   "KIA:6, PIN:3-5","0R","",["1R","2R","3R","4","1R","6","2","4","3R","6"], "Play as Discard only: ignore color of RNC / RPC determines target in multi card group / SNIPER CHECK: Action: RNC > Sniper RNC"),
    (157,"BRUSH",    "-1",   "1R", "",    ["1R","1","1","1","2R","1","3","5","4","7"], "OVR OT AFV 6 IM / 5 EL Others 6 IM / INF from: 2 shifts left"),
    (158,"MOVEMENT", "+1/⊙-1/-1","4","COWER",["1R","2R","2R","2","3R","2","4R","6","5","8"], "Retrograde NA: INF from/into 2 shifts right / FORD: Move sideways with RNC: Move removes Wire / OVR: Move sideways RR 5: Effect & RNC (no color)"),
    (159,"RALLY 1",  "+1",  "3",  "COWER",["1R","1R","3","4R","3R","4","5R","7","6","9"], "Retrograde NA: INF from/into 2 shifts right / FORD: Move sideways with RNC: Move removes Wire / OVR: Move sideways RR 5: Effect & RNC (no color)"),
    (160,"RALLY ALL","",    "0R", "",     ["1R","2R","1R","4","5","4","6R","8","7","0"], "Limit: one Rally card per Group per turn / Rallys 2 adj. groups with 1 unpinned Leader / Rallys 2 groups if each has unpinned Leader"),
    (161,"CONCEALED","+1/-1","3","COWER",["1R","1","2R","1","1R","5","7","1","8","1"], "Retrograde NA if RR to closest enemy will be < 0 / FORD: Move sideways with RNC: Move removes Wire / OVR: Move sideways RR 5: Effect & RNC (no color)"),
    (162,"CONCEALED","+1/-1","4","COWER",["1R","2R","3R","2","2R","6","1","2","9","2"], "Retrograde NA if RR to closest enemy will be < 0 / FORD: Move sideways with RNC: Move removes Wire / OVR: Move sideways RR 5: Effect & RNC (no color)"),
]

# --- xlsx 書き出し ---
HEADERS = ["#", "種別", "修正", "射程", "サブ",
           "1", "2b", "3", "4c", "5o", "6", "7w", "8", "9", "0r",
           "備考"]

def color_for(v):
    """値が '9R' のようなら赤、そうでなければ黒。"""
    if isinstance(v, str) and v.endswith("R"):
        return v[:-1], Font(color="CC0000", bold=True)
    return v, Font(color="000000")

def write_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "地形カード"
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")
        c.alignment = Alignment(horizontal="center")

    for card in CARDS:
        num, title, mod, rng, sub, vals, notes = card
        row = [num, title, mod, None, sub]
        for v in vals:
            row.append(None)
        row.append(notes)
        ws.append(row)
        r = ws.max_row
        # 射程セル(col 4)
        rv, rf = color_for(rng)
        ws.cell(row=r, column=4, value=rv).font = rf
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
        # 値セル(col 6..15)
        for i, v in enumerate(vals):
            vv, vf = color_for(v)
            cell = ws.cell(row=r, column=6 + i, value=vv)
            cell.font = vf
            cell.alignment = Alignment(horizontal="center")

    # 列幅
    widths = {1: 5, 2: 12, 3: 6, 4: 6, 5: 7,
              6: 4, 7: 4, 8: 4, 9: 4, 10: 4, 11: 4, 12: 4, 13: 4, 14: 4, 15: 4,
              16: 60}
    for col, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    out = "/Users/baron/Desktop/aigame/アップフロントVer2/upfrontimages/地形カード/terrain_cards.xlsx"
    wb.save(out)
    print(f"saved: {out}  ({len(CARDS)} cards)")

if __name__ == "__main__":
    write_xlsx()
